"""
Shared report generator for CUCM, MS Teams, and Webex Calling test results.
Exports: Excel (.xlsx), CSV, HTML
"""

import csv
from datetime import datetime
from pathlib import Path

RESULT_ANSWERED  = "ANSWERED"
RESULT_NO_ANSWER = "NO-ANSWER"
RESULT_BUSY      = "BUSY"
RESULT_REJECTED  = "REJECTED"
RESULT_ERROR     = "ERROR"

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False

COLOUR_MAP_HEX = {
    RESULT_ANSWERED:  "22C997",
    RESULT_NO_ANSWER: "F97316",
    RESULT_BUSY:      "F5C518",
    RESULT_REJECTED:  "F05252",
    RESULT_ERROR:     "A78BFA",
}

COLOUR_MAP_CSS = {
    RESULT_ANSWERED:  "#22c997",
    RESULT_NO_ANSWER: "#f97316",
    RESULT_BUSY:      "#f5c518",
    RESULT_REJECTED:  "#f05252",
    RESULT_ERROR:     "#a78bfa",
}


def _counts(rows: list[dict]) -> dict:
    return {
        "total":    len(rows),
        "answered": sum(1 for r in rows if r["result"] == RESULT_ANSWERED),
        "no_ans":   sum(1 for r in rows if r["result"] == RESULT_NO_ANSWER),
        "busy":     sum(1 for r in rows if r["result"] == RESULT_BUSY),
        "rejected": sum(1 for r in rows if r["result"] == RESULT_REJECTED),
        "errors":   sum(1 for r in rows if r["result"] == RESULT_ERROR),
    }


# ── CSV ───────────────────────────────────────────────────────────────────────

def export_csv(rows: list[dict], path: str):
    fields = ["#", "platform", "number", "result", "api_code",
              "duration_s", "started_at", "note"]
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        for i, r in enumerate(rows, 1):
            w.writerow({"#": i, **r})


# ── HTML ──────────────────────────────────────────────────────────────────────

def export_html(rows: list[dict], path: str, meta: dict):
    c = _counts(rows)
    total  = c["total"]
    pct_ok = round(c["answered"] / total * 100, 1) if total else 0
    platform = meta.get("platform", "")

    platform_badge_colour = {
        "MS Teams":      "#5B9CF6",
        "Webex Calling": "#00C86F",
        "CUCM":          "#F5A623",
    }.get(platform, "#6B7899")

    rows_html = ""
    for i, r in enumerate(rows, 1):
        col = COLOUR_MAP_CSS.get(r["result"], "#fff")
        rows_html += f"""<tr>
          <td>{i}</td>
          <td><span style="color:{platform_badge_colour};font-size:9px;
            border:1px solid {platform_badge_colour};padding:1px 5px">
            {r.get('platform', platform)}</span></td>
          <td><b>{r['number']}</b></td>
          <td style="color:{col};font-weight:700">{r['result']}</td>
          <td style="color:#6b7899">{r.get('api_code','')}</td>
          <td>{r.get('duration_s','')}</td>
          <td style="color:#6b7899;font-size:11px">{r.get('started_at','')}</td>
          <td style="color:#6b7899">{r.get('note','')}</td>
        </tr>"""

    server_info = ""
    if platform == "MS Teams":
        server_info = f"Tenant: {meta.get('tenant_id','')}"
    elif platform == "Webex Calling":
        server_info = f"User: {meta.get('user_display','')}"
    else:
        server_info = f"Server: {meta.get('server','')}:{meta.get('port','')}"

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>{platform} Call Test Report</title>
<style>
*{{box-sizing:border-box;margin:0;padding:0}}
body{{background:#0f1117;color:#dce3f0;font-family:'Courier New',monospace;font-size:12px;padding:28px}}
h1{{font-size:20px;letter-spacing:5px;color:{platform_badge_colour};margin-bottom:3px}}
.sub{{color:#6b7899;font-size:10px;letter-spacing:2px;margin-bottom:24px}}
.stats{{display:flex;gap:10px;flex-wrap:wrap;margin-bottom:24px}}
.stat{{background:#1e2230;border:1px solid #2a2f3f;padding:14px 18px;min-width:105px}}
.sv{{font-size:26px;font-weight:700;margin-bottom:2px}}
.sl{{font-size:8px;letter-spacing:2px;text-transform:uppercase;color:#6b7899}}
.s1{{color:{platform_badge_colour}}}.s2{{color:#22c997}}.s3{{color:#f97316}}
.s4{{color:#f5c518}}.s5{{color:#f05252}}.s6{{color:#a78bfa}}
table{{width:100%;border-collapse:collapse}}
th{{background:#1e2230;padding:8px 12px;text-align:left;font-size:8px;
    letter-spacing:1.5px;text-transform:uppercase;color:#6b7899;
    border-bottom:1px solid #2a2f3f;position:sticky;top:0}}
td{{padding:7px 12px;border-bottom:1px solid rgba(255,255,255,.04)}}
tr:hover td{{background:rgba(255,255,255,.02)}}
</style>
</head>
<body>
<h1>{platform.upper()}  CALL TEST REPORT</h1>
<div class="sub">Generated {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} &nbsp;|&nbsp; {server_info}</div>
<div class="stats">
  <div class="stat"><div class="sv s1">{total}</div><div class="sl">Total</div></div>
  <div class="stat"><div class="sv s2">{c['answered']}</div><div class="sl">Answered</div></div>
  <div class="stat"><div class="sv s3">{c['no_ans']}</div><div class="sl">No Answer</div></div>
  <div class="stat"><div class="sv s4">{c['busy']}</div><div class="sl">Busy</div></div>
  <div class="stat"><div class="sv s5">{c['rejected']}</div><div class="sl">Rejected</div></div>
  <div class="stat"><div class="sv s6">{c['errors']}</div><div class="sl">Error</div></div>
  <div class="stat"><div class="sv s2">{pct_ok}%</div><div class="sl">Success</div></div>
</div>
<table>
<thead><tr>
  <th>#</th><th>Platform</th><th>Number / Target</th><th>Result</th>
  <th>Code</th><th>Duration (s)</th><th>Timestamp</th><th>Notes</th>
</tr></thead>
<tbody>{rows_html}</tbody>
</table>
</body></html>"""

    with open(path, "w", encoding="utf-8") as f:
        f.write(html)


# ── Excel ─────────────────────────────────────────────────────────────────────

def export_xlsx(rows: list[dict], path: str, meta: dict):
    if not HAS_XLSX:
        raise RuntimeError("openpyxl not installed — run: pip install openpyxl")

    platform = meta.get("platform", "Call Test")
    accent   = {
        "MS Teams":      "5B9CF6",
        "Webex Calling": "00C86F",
        "CUCM":          "F5A623",
    }.get(platform, "3D8EF0")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    def fill(c): return PatternFill("solid", fgColor=c)
    def font(c="DCE3F0", bold=False, sz=11):
        return Font(color=c, bold=bold, size=sz, name="Courier New")
    def bdr():
        s = Side(style="thin", color="2A2F3F")
        return Border(left=s, right=s, top=s, bottom=s)

    ctr = Alignment(horizontal="center", vertical="center")
    lft = Alignment(horizontal="left",   vertical="center")

    # Title
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value     = f"{platform.upper()}  —  CALL TEST REPORT"
    c.fill      = fill("0F1117"); c.font = Font(color=accent, bold=True, size=15, name="Courier New")
    c.alignment = ctr
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    c = ws["A2"]
    server_info = (f"Tenant: {meta.get('tenant_id','')}" if platform == "MS Teams"
                   else f"User: {meta.get('user_display','')}" if platform == "Webex Calling"
                   else f"Server: {meta.get('server','')}:{meta.get('port','')}")
    c.value     = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}   |   {server_info}"
    c.fill      = fill("0F1117"); c.font = font("6B7899", sz=9); c.alignment = ctr
    ws.row_dimensions[2].height = 16

    # Summary
    cnts = _counts(rows)
    pct  = round(cnts["answered"] / cnts["total"] * 100, 1) if cnts["total"] else 0
    summary = [
        ("TOTAL",     cnts["total"],    accent),
        ("ANSWERED",  cnts["answered"], "22C997"),
        ("NO-ANSWER", cnts["no_ans"],   "F97316"),
        ("BUSY",      cnts["busy"],     "F5C518"),
        ("REJECTED",  cnts["rejected"], "F05252"),
        ("ERROR",     cnts["errors"],   "A78BFA"),
        ("SUCCESS %", f"{pct}%",        "22C997"),
    ]
    for ci, (lbl, val, col) in enumerate(summary, 1):
        lc = ws.cell(row=3, column=ci, value=lbl)
        vc = ws.cell(row=4, column=ci, value=val)
        lc.fill = fill("1E2230"); lc.font = font("6B7899", sz=8); lc.alignment = ctr
        vc.fill = fill("1E2230"); vc.font = Font(color=col, bold=True, size=18, name="Courier New")
        vc.alignment = ctr
    ws.row_dimensions[3].height = 14; ws.row_dimensions[4].height = 32

    # Header row
    headers = ["#", "Platform", "Number / Target", "Result", "Code", "Duration (s)", "Timestamp", "Notes"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=6, column=ci, value=h)
        c.fill = fill("181C24"); c.font = font("6B7899", bold=True, sz=9)
        c.alignment = ctr; c.border = bdr()
    ws.row_dimensions[6].height = 18

    # Data
    for ri, row in enumerate(rows, 7):
        col = COLOUR_MAP_HEX.get(row["result"], "DCE3F0")
        bg  = "181C24" if ri % 2 == 0 else "1E2230"
        vals = [ri - 6, row.get("platform", platform), row["number"],
                row["result"], row.get("api_code",""),
                row.get("duration_s",""), row.get("started_at",""), row.get("note","")]
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill   = fill(bg); c.border = bdr()
            c.alignment = lft if ci in (3, 8) else ctr
            if ci == 4:
                c.font = Font(color=col, bold=True, size=11, name="Courier New")
            elif ci == 3:
                c.font = font(bold=True)
            else:
                c.font = font("6B7899")

    widths = [4, 14, 22, 13, 9, 13, 20, 40]
    for ci, w in enumerate(widths, 1):
        from openpyxl.utils import get_column_letter; ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes   = "A7"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = accent

    # Raw sheet
    ws2 = wb.create_sheet("Raw Data")
    ws2.append(["#","platform","number","result","api_code","duration_s","started_at","note"])
    for i, r in enumerate(rows, 1):
        ws2.append([i, r.get("platform",""), r["number"], r["result"],
                    r.get("api_code",""), r.get("duration_s",""),
                    r.get("started_at",""), r.get("note","")])

    wb.save(path)
