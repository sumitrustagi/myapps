from flask import Flask, request, render_template, send_file
import re
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from datetime import datetime

app = Flask(__name__)

# ─── Styles (defined once) ───────────────────────────────────────────────────
HDR_FONT    = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
HDR_FILL    = PatternFill('solid', fgColor='01696F')
TITLE_FONT  = Font(name='Calibri', bold=True, size=14, color='01696F')
SUB_FONT    = Font(name='Calibri', italic=True, size=10, color='6B6962')
DATA_FONT   = Font(name='Calibri', size=11)
BOLD_FONT   = Font(name='Calibri', bold=True, size=11)
CENTER      = Alignment(horizontal='center', vertical='center')
LEFT        = Alignment(horizontal='left', vertical='center', indent=1)
RIGHT       = Alignment(horizontal='right', vertical='center')
THIN        = Border(
    left=Side(style='thin', color='D4D1CA'),
    right=Side(style='thin', color='D4D1CA'),
    top=Side(style='thin', color='D4D1CA'),
    bottom=Side(style='thin', color='D4D1CA'),
)

# ─── Conversion helpers ──────────────────────────────────────────────────────
def target_capacity(model):
    return {"VG420-84": 84, "VG420-144": 144, "VG410-24": 24, "VG410-48": 48}.get(model, 84)

def flatten_port(legacy, base_slot, ports_per_module):
    m = re.fullmatch(r"(\d+)/(\d+)/(\d+)", legacy.strip())
    if not m:
        return None
    slot_a, slot_b, port = map(int, m.groups())
    module_slot = slot_b if slot_a == base_slot else slot_a
    flat = (module_slot - base_slot) * ports_per_module + port
    return f"0/0/{flat}" if flat >= 0 else None

def replace_hostname(cfg, new_hostname):
    if not new_hostname.strip():
        return cfg
    if re.search(r"(?m)^hostname\s+\S+", cfg):
        return re.sub(r"(?m)^hostname\s+\S+", f"hostname {new_hostname.strip()}", cfg)
    return f"hostname {new_hostname.strip()}\n" + cfg

def replace_ip(cfg, interface_name, new_ip, new_mask, warnings):
    if not (new_ip.strip() and new_mask.strip() and interface_name.strip()):
        return cfg
    pattern = rf"(?ms)(^interface\s+{re.escape(interface_name.strip())}\n)(.*?)(?=^!|^interface\s|^router\s|^voice\s|^dial-peer\s|^line\s|\Z)"
    m = re.search(pattern, cfg)
    if not m:
        warnings.append(f'Interface "{interface_name}" not found — management IP was not replaced.')
        return cfg
    full_match = m.group(0)
    if re.search(r"(?m)^ ip address\s+\S+\s+\S+", full_match):
        new_block = re.sub(r"(?m)^ ip address\s+\S+\s+\S+",
                           f" ip address {new_ip.strip()} {new_mask.strip()}", full_match)
    else:
        lines = full_match.splitlines()
        lines.insert(1, f" ip address {new_ip.strip()} {new_mask.strip()}")
        new_block = "\n".join(lines)
    return cfg.replace(full_match, new_block, 1)

# ─── Extract port→extension mapping ─────────────────────────────────────────
def extract_port_extension_map(cfg):
    """
    Returns a list of (port, extension) tuples found in dial-peer + voice-port blocks.
    Searches for:
      dial-peer voice <tag> pots
        destination-pattern <ext>
        port <X/Y/Z>
    """
    mapping = []
    # Pattern: dial-peer block with destination-pattern and port
    dp_blocks = re.findall(
        r'dial-peer voice \d+ pots.*?(?=dial-peer|\Z)', cfg, re.DOTALL | re.IGNORECASE
    )
    for block in dp_blocks:
        port_m = re.search(r'port\s+(\S+)', block)
        ext_m  = re.search(r'destination-pattern\s+(\S+)', block)
        if port_m and ext_m:
            port = port_m.group(1).strip()
            ext  = re.sub(r'[^0-9+*#]', '', ext_m.group(1))
            if port and ext:
                mapping.append((port, ext))
    # Remove duplicates, sort by port
    seen = set()
    result = []
    for p, e in mapping:
        if p not in seen:
            seen.add(p)
            result.append((p, e))
    result.sort(key=lambda x: [int(n) for n in x[0].split('/') if n.isdigit()])
    return result

# ─── Build port mapping table ────────────────────────────────────────────────
def build_port_mapping(old_cfg, new_cfg, base_slot, ports_per_module):
    old_map = {p: e for p, e in extract_port_extension_map(old_cfg)}
    new_map = {p: e for p, e in extract_port_extension_map(new_cfg)}
    return old_map, new_map

# ─── Main conversion ─────────────────────────────────────────────────────────
def convert_config(cfg, target_model, base_slot, ports_per_module,
                   new_hostname, new_ip, new_mask, interface_name):
    warnings = []
    cap = target_capacity(target_model)
    found_ports = sorted(set(re.findall(r"(?<!\d)(\d+/\d+/\d+)(?!\d)", cfg)),
                         key=lambda x: list(map(int, x.split("/"))))
    mapping = {}
    overflow = []
    for p in found_ports:
        mapped = flatten_port(p, base_slot, ports_per_module)
        if mapped is None:
            continue
        idx = int(mapped.split("/")[-1])
        if idx >= cap:
            overflow.append((p, mapped))
        mapping[p] = mapped

    new_cfg = cfg
    for old, new in sorted(mapping.items(), key=lambda x: len(x[0]), reverse=True):
        new_cfg = re.sub(rf"(?<!\d){re.escape(old)}(?!\d)", new, new_cfg)
    new_cfg = replace_hostname(new_cfg, new_hostname)
    new_cfg = replace_ip(new_cfg, interface_name, new_ip, new_mask, warnings)

    banner = [
        "!",
        f"! Converted for         : {target_model}",
        f"! Legacy base slot      : {base_slot}",
        f"! Legacy ports/module   : {ports_per_module}",
        f"! Total ports remapped  : {len(mapping)}",
        "! *** Review voice-port, dial-peer and hardware module commands before production use ***",
        "!",
    ]
    if overflow:
        warnings.append(f"{len(overflow)} port(s) exceed the {target_model} capacity of {cap} ports.")
        warnings.append("Overflow: " + ", ".join(f"{o}->{n}" for o, n in overflow[:10]))
    if not mapping:
        warnings.append("No legacy slot/port references detected in X/Y/Z format.")
    if "sm-d-" in new_cfg.lower():
        warnings.append("Legacy SM-D module lines detected — remove hardware-specific commands not supported on VG410/VG420.")

    result_text = "\n".join(banner) + "\n" + new_cfg.strip() + "\n"
    return result_text, warnings, len(mapping), mapping, cfg, new_cfg

# ─── Build XLSX workbook ──────────────────────────────────────────────────────
def build_xlsx(old_cfg, new_cfg, old_hostname, new_hostname, target_model):
    old_map = extract_port_extension_map(old_cfg)
    new_map = extract_port_extension_map(new_cfg)

    wb = Workbook()

    def make_sheet(wb, title, rows, sheet_label, hostname, is_first=False):
        ws = wb.active if is_first else wb.create_sheet(title)
        ws.title = title
        ws.sheet_view.showGridLines = False
        ws.column_dimensions['A'].width = 3

        # Title
        ws.merge_cells('B2:D2')
        c = ws['B2']
        c.value = f"{sheet_label}  |  {hostname or 'N/A'}"
        c.font = TITLE_FONT
        c.alignment = LEFT
        ws.row_dimensions[2].height = 28

        # Subtitle
        ws.merge_cells('B3:D3')
        s = ws['B3']
        s.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  {len(rows)} port(s)"
        s.font = SUB_FONT
        s.alignment = LEFT
        ws.row_dimensions[3].height = 16

        # Header row
        headers = ['Port Number', 'Extension', 'Notes']
        for ci, h in enumerate(headers, start=2):
            cell = ws.cell(row=5, column=ci, value=h)
            cell.font = HDR_FONT
            cell.fill = HDR_FILL
            cell.alignment = CENTER
            cell.border = THIN
        ws.row_dimensions[5].height = 22

        # Data rows
        sorted_rows = sorted(rows, key=lambda x: [int(n) for n in x[0].split('/') if n.isdigit()])
        for ri, (port, ext) in enumerate(sorted_rows, start=6):
            ws.cell(row=ri, column=2, value=port).alignment = LEFT
            ws.cell(row=ri, column=3, value=ext).alignment = CENTER
            ws.cell(row=ri, column=4, value='').alignment = LEFT
            for ci in range(2, 5):
                ws.cell(row=ri, column=ci).font = DATA_FONT
                ws.cell(row=ri, column=ci).border = THIN
            ws.row_dimensions[ri].height = 18

        last_data_row = 5 + len(sorted_rows)

        # Excel Table
        if sorted_rows:
            tbl = Table(
                displayName=f"tbl_{title.replace(' ', '_')}",
                ref=f"B5:D{last_data_row}"
            )
            tbl.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2", showFirstColumn=False,
                showLastColumn=False, showRowStripes=True, showColumnStripes=False
            )
            ws.add_table(tbl)

        # Column widths
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 30

        # Freeze panes
        ws.freeze_panes = 'B6'

        # Footer
        footer_row = last_data_row + 2
        ws.merge_cells(f'B{footer_row}:D{footer_row}')
        f = ws.cell(row=footer_row, column=2,
                    value=f"Source: Cisco VG Config Converter  |  {datetime.now().strftime('%Y-%m-%d')}")
        f.font = Font(name='Calibri', size=9, color='B0AFA9', italic=True)
        f.alignment = LEFT

        return ws

    make_sheet(wb, 'OLD_Config', old_map, 'Old Config (VG350)', old_hostname, is_first=True)
    make_sheet(wb, 'NEW_Config', new_map, f'New Config ({target_model})', new_hostname, is_first=False)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ─── Routes ──────────────────────────────────────────────────────────────────
@app.route("/", methods=["GET"])
def index():
    return render_template("index.html", result=None, warnings=[], result_hidden="",
                           port_count=0, show_xlsx=False)

@app.route("/convert", methods=["POST"])
def convert():
    uploaded = request.files.get("config_file")
    text = request.form.get("config_text", "") or ""
    if uploaded and uploaded.filename:
        text = uploaded.read().decode("utf-8", errors="ignore")
    if not text.strip():
        return render_template("index.html", result=None,
                               warnings=["No config provided. Upload a file or paste a running-config."],
                               result_hidden="", port_count=0, show_xlsx=False)

    target_model   = request.form.get("target_model", "VG420-84")
    base_slot      = int(request.form.get("old_base_slot", "0") or 0)
    ports_per_mod  = int(request.form.get("old_ports_per_module", "24") or 24)
    new_hostname   = request.form.get("new_hostname", "")
    new_ip         = request.form.get("new_ip", "")
    new_mask       = request.form.get("new_mask", "255.255.255.0")
    iface          = request.form.get("interface_name", "GigabitEthernet0/0/0")
    old_hostname   = request.form.get("old_hostname_hint", "VG350")

    result, warnings, port_count, mapping, old_cfg, new_cfg = convert_config(
        text, target_model, base_slot, ports_per_mod, new_hostname, new_ip, new_mask, iface
    )

    # stash old_cfg in a hidden field for xlsx generation
    return render_template("index.html", result=result, warnings=warnings,
                           result_hidden=result, port_count=port_count, show_xlsx=True,
                           old_cfg_hidden=old_cfg, new_cfg_hidden=new_cfg,
                           old_hostname=old_hostname, new_hostname=new_hostname,
                           target_model=target_model)

@app.route("/download", methods=["POST"])
def download():
    content = request.form.get("content", "")
    return send_file(BytesIO(content.encode("utf-8")), as_attachment=True,
                     download_name="converted-vg-config.txt", mimetype="text/plain")

@app.route("/download_xlsx", methods=["POST"])
def download_xlsx():
    old_cfg      = request.form.get("old_cfg", "")
    new_cfg      = request.form.get("new_cfg", "")
    old_hostname = request.form.get("old_hostname", "VG350")
    new_hostname = request.form.get("new_hostname", "VG4xx")
    target_model = request.form.get("target_model", "VG420-84")
    buf = build_xlsx(old_cfg, new_cfg, old_hostname, new_hostname, target_model)
    return send_file(buf, as_attachment=True,
                     download_name="vg-port-extension-map.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)